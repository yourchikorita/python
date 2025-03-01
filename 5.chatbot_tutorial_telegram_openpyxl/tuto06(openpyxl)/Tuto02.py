# -*- coding: utf-8 -*-
"""
Telegram bot 챗봇에 사용자가 처음 들어왔을때 반겨주는
Welcome message 를 작성해 봅시다.
"""
from openpyxl import load_workbook

EXCEL_FILE_NAME = "DB.xlsx"
db = load_workbook(filename=EXCEL_FILE_NAME)
user_db = db['User_DB']


def find_user_row(user_id, user_name):
    # user_exist : 사용자가 방문 경험이 있는지 유무(있으면 True, 없으면 False)
    # user_exist == False , is_welcome == True 
    user_exist = False
    is_welcome = False
            
    
    for row in user_db.rows:
        if row[0].value == user_id:
            user_exist = True
            user_row = row[0].row
            
        

    if not user_exist:
        is_welcome = True
        # max_row 는 행의 최대값을 리턴해주는데,
        # 한번이라도 데이터가 적힌 행 자체를 삭제해줘야 빈칸이 안생김
        user_db[user_db.max_row + 1][0].value = user_id
        user_db[user_db.max_row][1].value = user_name
        user_db[user_db.max_row][2].value = 0
        user_row = user_db.max_row
        
    
    #user_db.cell(row=user_row, column=3).value = user_db.cell(row=1, column=1).value + 1
   # user_db.cell(row=user_row, column=3).value += 1 
    db.save(EXCEL_FILE_NAME)
    
    return user_row, is_welcome

# 실습 2
# 사용자가 말을 몇번 걸었는지 엑셀파일에 누적하는 함수를 만들어보기
def user_counter(user_row):
    user_db.cell(row=user_row, column=3).value += 1 
    db.save(EXCEL_FILE_NAME)
